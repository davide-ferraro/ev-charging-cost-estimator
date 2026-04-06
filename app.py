import streamlit as st
import pandas as pd
import plotly.graph_objs as go
from collections import defaultdict
from io import BytesIO
from datetime import date
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.styles import Font
from utils.calls import (
    medium_requirement, hard_requirement, case_definition,
    compute_all_costs, cost_breakdown_calculation,
    cost_breakdown_sensitivity_calculation,
    cost_breakdown_double_sensitivity_calculation
)

# ---------------------------------------------------------------------------
# Theme constants – light theme
# ---------------------------------------------------------------------------
ELECTRIC_BLUE = "#0066FF"
LIME_GREEN = "#22C55E"
BG = "#ffffff"
BG_CARD = "#f7f8fa"
BORDER = "#e2e6ec"
TEXT_PRIMARY = "#0a0a0a"
TEXT_SECONDARY = "#5a6270"
TEXT_MUTED = "#9ca3af"

# Plotly chart palette – electric blue & lime green anchored
CHART_COLORS = [
    ELECTRIC_BLUE, LIME_GREEN, "#0099FF", "#34D399", "#1E90FF",
    "#4ADE80", "#4169E1", "#10B981", "#6495ED", "#6EE7B7",
    "#3B82F6", "#86EFAC", "#5F9EA0", "#00C49A",
]

st.set_page_config(
    page_title="EV Charging Cost Estimator",
    page_icon="\u26a1",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Custom CSS – light theme, full-width, big controls
# ---------------------------------------------------------------------------
st.markdown(f"""
<style>
    /* ---- Global ---- */
    .stApp {{
        background-color: {BG};
        color: {TEXT_PRIMARY};
    }}
    header[data-testid="stHeader"] {{
        background-color: {BG};
    }}

    .block-container {{
        max-width: 1400px;
        padding: 2rem 3rem 4rem 3rem;
    }}

    /* ---- Typography ---- */
    .hero-title {{
        font-size: 2.8rem;
        font-weight: 900;
        letter-spacing: -0.04em;
        background: linear-gradient(135deg, {ELECTRIC_BLUE}, {LIME_GREEN});
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.15rem;
    }}
    .hero-sub {{
        font-size: 1.1rem;
        color: {TEXT_SECONDARY};
        margin-bottom: 2.5rem;
    }}

    .section-number {{
        display: inline-block;
        background: {ELECTRIC_BLUE};
        color: #fff;
        font-weight: 800;
        font-size: 0.85rem;
        width: 28px;
        height: 28px;
        line-height: 28px;
        text-align: center;
        border-radius: 8px;
        margin-right: 10px;
        vertical-align: middle;
    }}
    .section-heading {{
        font-size: 1.5rem;
        font-weight: 800;
        color: {TEXT_PRIMARY};
        letter-spacing: -0.02em;
        margin-bottom: 0.15rem;
    }}
    .section-desc {{
        font-size: 0.95rem;
        color: {TEXT_SECONDARY};
        margin-bottom: 1.5rem;
    }}

    .group-label {{
        font-size: 1.05rem;
        font-weight: 700;
        color: {ELECTRIC_BLUE};
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin: 1.8rem 0 0.6rem 0;
        padding-bottom: 0.4rem;
        border-bottom: 2px solid {ELECTRIC_BLUE}22;
    }}

    .disabled-banner {{
        background: {BG_CARD};
        color: {TEXT_MUTED};
        border: 1px dashed {BORDER};
        border-radius: 10px;
        padding: 0.7rem 1rem;
        font-size: 0.9rem;
        text-align: center;
        margin-bottom: 0.75rem;
    }}

    /* ---- Results section light-green background ---- */
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.results-marker) {{
        background: rgba(34, 197, 94, 0.045);
        border: 1px solid rgba(34, 197, 94, 0.13);
        border-radius: 16px;
        padding: 1.5rem !important;
    }}

    /* ---- Dividers ---- */
    .section-divider {{
        border: none;
        border-top: 2px solid #eef0f4;
        margin: 2.5rem 0 2rem 0;
    }}

    /* ---- Cards for analysis options ---- */
    .option-card {{
        background: {BG_CARD};
        border: 1px solid {BORDER};
        border-radius: 14px;
        padding: 1.4rem 1.6rem;
        margin-bottom: 0.5rem;
        transition: border-color 0.2s;
    }}
    .option-card:hover {{
        border-color: {ELECTRIC_BLUE}66;
    }}
    .option-card h4 {{
        margin: 0 0 0.3rem 0;
        font-size: 1.1rem;
        font-weight: 700;
        color: {TEXT_PRIMARY};
    }}
    .option-card p {{
        margin: 0;
        font-size: 0.88rem;
        color: {TEXT_SECONDARY};
        line-height: 1.45;
    }}

    /* ---- Number inputs ---- */
    /* Fix clipping: all wrappers must allow overflow */
    div[data-testid="stNumberInput"],
    div[data-testid="stNumberInput"] > div,
    div[data-testid="stNumberInput"] > div > div {{
        overflow: visible !important;
    }}

    div[data-testid="stNumberInput"] input {{
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        border-radius: 10px !important;
        background: {BG} !important;
        color: {TEXT_PRIMARY} !important;
        border: 2px solid {BORDER} !important;
        text-align: center !important;
    }}
    div[data-testid="stNumberInput"] input:focus {{
        border-color: {ELECTRIC_BLUE} !important;
        box-shadow: 0 0 0 3px {ELECTRIC_BLUE}18 !important;
    }}
    div[data-testid="stNumberInput"] button {{
        border-radius: 8px !important;
        background: {BG_CARD} !important;
        color: {ELECTRIC_BLUE} !important;
        border: 2px solid {BORDER} !important;
    }}
    div[data-testid="stNumberInput"] button:hover {{
        background: {ELECTRIC_BLUE}12 !important;
        border-color: {ELECTRIC_BLUE} !important;
    }}

    /* Labels */
    div[data-testid="stNumberInput"] label p {{
        font-size: 0.88rem !important;
        color: {TEXT_SECONDARY} !important;
        font-weight: 600 !important;
    }}
    /* Help icon (?) – small, inline next to label */
    div[data-testid="stNumberInput"] label div[data-testid="stTooltipIcon"] svg {{
        width: 14px !important;
        height: 14px !important;
    }}

    /* ---- Radio buttons ---- */
    div[data-testid="stRadio"] label span {{
        font-size: 1rem !important;
    }}
    div[data-testid="stRadio"] div[role="radiogroup"] label {{
        background: {BG} !important;
        border: 2px solid {BORDER} !important;
        border-radius: 10px !important;
        padding: 0.5rem 1.2rem !important;
        margin-right: 0.5rem !important;
        transition: all 0.15s !important;
    }}
    div[data-testid="stRadio"] div[role="radiogroup"] label[data-checked="true"] {{
        background: {ELECTRIC_BLUE}0d !important;
        border-color: {ELECTRIC_BLUE} !important;
    }}

    /* ---- Checkboxes ---- */
    div[data-testid="stCheckbox"] label span {{
        font-size: 1rem !important;
    }}

    /* ---- Selectbox / Multiselect ---- */
    div[data-testid="stSelectbox"] > div > div,
    div[data-testid="stMultiSelect"] > div > div {{
        background: {BG} !important;
        border: 2px solid {BORDER} !important;
        border-radius: 12px !important;
        font-size: 1rem !important;
        min-height: 3rem !important;
    }}

    /* ---- Tabs ---- */
    button[data-baseweb="tab"] {{
        font-size: 1.05rem !important;
        font-weight: 600 !important;
        color: {TEXT_SECONDARY} !important;
        padding: 0.8rem 1.4rem !important;
    }}
    button[data-baseweb="tab"][aria-selected="true"] {{
        color: {ELECTRIC_BLUE} !important;
    }}

    /* ---- Primary button ---- */
    button[data-testid="stBaseButton-primary"] {{
        background: linear-gradient(135deg, {ELECTRIC_BLUE}, {LIME_GREEN}) !important;
        color: #fff !important;
        font-size: 1.2rem !important;
        font-weight: 800 !important;
        height: 3.6rem !important;
        border-radius: 14px !important;
        border: none !important;
        letter-spacing: 0.02em;
        transition: opacity 0.2s;
    }}
    button[data-testid="stBaseButton-primary"]:hover {{
        opacity: 0.9;
    }}

    /* ---- Download button ---- */
    button[data-testid="stDownloadButton"] > button,
    div[data-testid="stDownloadButton"] button {{
        background: {BG} !important;
        border: 2px solid {LIME_GREEN} !important;
        color: {LIME_GREEN} !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
    }}

    /* ---- Dataframe ---- */
    div[data-testid="stDataFrame"] {{
        border-radius: 12px;
        overflow: hidden;
    }}

    /* ---- Expander ---- */
    details[data-testid="stExpander"] {{
        background: {BG_CARD} !important;
        border: 1px solid {BORDER} !important;
        border-radius: 14px !important;
    }}
    details[data-testid="stExpander"] summary span {{
        font-size: 1rem !important;
        font-weight: 600 !important;
    }}

    /* ---- Error banner ---- */
    div[data-testid="stAlert"] {{
        border-radius: 12px !important;
    }}
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Plotly shared layout template – light theme
# ---------------------------------------------------------------------------
PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor=BG_CARD,
    font=dict(family="Inter, system-ui, sans-serif", color=TEXT_PRIMARY, size=13),
    title_font=dict(size=18, color=TEXT_PRIMARY),
    legend=dict(
        bgcolor="rgba(0,0,0,0)",
        font=dict(color=TEXT_SECONDARY, size=12),
    ),
    xaxis=dict(
        gridcolor="#e8eaed", zerolinecolor="#d0d4da",
        title_font=dict(color=TEXT_SECONDARY),
        tickfont=dict(color=TEXT_SECONDARY),
    ),
    yaxis=dict(
        gridcolor="#e8eaed", zerolinecolor="#d0d4da",
        title_font=dict(color=TEXT_SECONDARY),
        tickfont=dict(color=TEXT_SECONDARY),
    ),
    margin=dict(l=60, r=30, t=60, b=50),
    hoverlabel=dict(
        bgcolor="#ffffff",
        font_size=13,
        font_color=TEXT_PRIMARY,
        bordercolor=ELECTRIC_BLUE,
    ),
)

# ---------------------------------------------------------------------------
# Input definitions (identical to original)
# ---------------------------------------------------------------------------
INPUT_DICT = {
    "Number of chargers": (1, 100, 1, 4, False),
    "Power of chargers [kW]": (10, 350, 50, 100, False),
    "Low Voltage level [V]": (200, 1000, 20, 480, False),
    "Grid Connection capacity [kVA]": (0, 5000, 100, 500, False),
    "Distance between the rectifier and the chargers [meters]": (0, 1000, 50, 100, False),
    "Load Power Factor": (0.70, 1.00, 0.05, 0.90, True),
    "Maximum power for Low Voltage Connections in your area [kVA]": (10, 1000, 10, 200, False),
    "Available Capacity of the closest MV/LV transformer [kVA]": (0, 5000, 100, 1000, False),
    "Distance between your premises and the closest MV/LV transformer [meters]": (0, 5000, 50, 500, False),
    "Medium Voltage level [kV]": (10, 45, 1, 10, False),
    "Transformer size safety margin for your new transformer [%]": (0, 100, 5, 20, False),
    "Distance between your premises and the closest Medium Voltage Access point [meters]": (0, 5000, 50, 500, False),
    "Land to prepare for hosting the parking lot [m^2]": (0, 100000, 100, 0, False),
}

INPUT_LABELS = list(INPUT_DICT.keys())

TOOLTIPS = {
    "Grid Connection capacity [kVA]": (
        "If you DON'T HAVE a transformer: max apparent power allowed. "
        "If you HAVE a transformer: enter spare capacity (e.g. 500 - 300 = 200 kVA)."
    ),
    "Maximum power for Low Voltage Connections in your area [kVA]": (
        "200 kVA is a reasonable default in Europe. If unsure, leave at 200."
    ),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def render_input(label, disabled=False, key_prefix=""):
    minv, maxv, step, default, is_float = INPUT_DICT[label]
    help_text = TOOLTIPS.get(label)
    key = f"{key_prefix}{label}" if key_prefix else label
    if is_float:
        return st.number_input(
            label, min_value=float(minv), max_value=float(maxv),
            value=float(default), step=float(step), help=help_text,
            disabled=disabled, key=key,
        )
    else:
        return st.number_input(
            label, min_value=int(minv), max_value=int(maxv),
            value=int(default), step=int(step), help=help_text,
            disabled=disabled, key=key,
        )


def breakdown_to_excel(data_dict):
    df = pd.DataFrame(list(data_dict.items()), columns=["Component", "Cost"])

    def parse_cost(val):
        if isinstance(val, str) and val.strip().startswith("\u20ac"):
            return float(val.replace("\u20ac", "").replace(",", ""))
        return None

    df["Cost"] = df["Cost"].apply(parse_cost)
    buf = BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        cost_cell = ws[f"B{row}"]
        component_cell = ws[f"A{row}"]
        if cost_cell.value is None:
            component_cell.font = Font(bold=True)
        else:
            cost_cell.number_format = '\u20ac#,##0.00'
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def sensitivity_to_excel(param, value_dict):
    cost_components = list(next(iter(value_dict.values())).keys())
    values = list(value_dict.keys())
    rows = []
    for component in cost_components:
        row = {"Component": component}
        for val in values:
            row[str(val)] = value_dict[val].get(component, "\u20ac0.00")
        rows.append(row)
    df = pd.DataFrame(rows)
    buf = BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf


REPORT_URL = (
    "https://www.diva-portal.org/smash/record.jsf?dswid=1523"
    "&pid=diva2%3A2007635&c=5&searchType=SIMPLE&language=en"
    "&query=investment+cost+charging+infrastructure+heavy+duty"
    "&af=%5B%5D&aq=%5B%5B%5D%5D&aq2=%5B%5B%5D%5D&aqe=%5B%5D"
    "&noOfRows=50&sortOrder=author_sort_asc&sortOrder2=title_sort_asc"
    "&onlyFullText=true&sf=all"
)


def _fig_to_bytes(fig, width=900, height=500):
    """Render a Plotly figure to PNG bytes via kaleido. Returns None if unavailable."""
    try:
        return fig.to_image(format="png", width=width, height=height, scale=2)
    except Exception:
        return None


def _parse_euro(val):
    """'€12,345.67' -> 12345.67  |  non-cost strings -> None"""
    if isinstance(val, str) and "\u20ac" in val:
        return float(val.replace("\u20ac", "").replace(",", ""))
    return None


import os as _os
_FONT_DIR = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "fonts")
_FONT = "DejaVu"


class _PDF(FPDF):
    """Thin FPDF subclass with header/footer branding (Unicode-safe)."""

    def __init__(self):
        super().__init__()
        self.add_font(_FONT, "", f"{_FONT_DIR}/DejaVuSans.ttf")
        self.add_font(_FONT, "B", f"{_FONT_DIR}/DejaVuSans-Bold.ttf")
        self.add_font(_FONT, "I", f"{_FONT_DIR}/DejaVuSans-Oblique.ttf")
        self.add_font(_FONT, "BI", f"{_FONT_DIR}/DejaVuSans-BoldOblique.ttf")

    def header(self):
        self.set_font(_FONT, "B", 10)
        self.set_text_color(100, 100, 100)
        self.cell(0, 8, "EV Charging Infrastructure  |  Investment Cost Estimation Report", align="R")
        self.ln(12)

    def footer(self):
        self.set_y(-15)
        self.set_font(_FONT, "I", 8)
        self.set_text_color(160, 160, 160)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

    # ---- convenience methods ----
    def section_title(self, num, text):
        self.set_font(_FONT, "B", 16)
        self.set_text_color(0, 102, 255)
        self.cell(0, 12, f"{num}.  {text}")
        self.ln(14)

    def sub_heading(self, text):
        self.set_font(_FONT, "B", 12)
        self.set_text_color(40, 40, 40)
        self.cell(0, 10, text)
        self.ln(11)

    def body_text(self, text):
        self.set_font(_FONT, "", 10)
        self.set_text_color(50, 50, 50)
        self.multi_cell(0, 6, text)
        self.ln(3)

    def key_value(self, key, value):
        """Render a label: value pair. The label wraps if needed."""
        usable = self.w - self.l_margin - self.r_margin
        val_w = 50
        key_w = usable - val_w
        y_before = self.get_y()
        x_start = self.l_margin
        # Key (may wrap)
        self.set_font(_FONT, "", 10)
        self.set_text_color(90, 90, 90)
        self.set_xy(x_start, y_before)
        self.multi_cell(key_w, 6, key)
        y_after_key = self.get_y()
        # Value (right-aligned, vertically centered)
        key_h = y_after_key - y_before
        self.set_font(_FONT, "B", 10)
        self.set_text_color(30, 30, 30)
        val_y = y_before + (key_h - 6) / 2
        self.set_xy(x_start + key_w, val_y)
        self.cell(val_w, 6, str(value), align="R")
        self.set_y(y_after_key + 2)

    def add_table(self, headers, rows, col_widths=None, font_size=9):
        """Render a table that auto-pages. col_widths are computed if None."""
        if col_widths is None:
            usable = self.w - self.l_margin - self.r_margin
            w = usable / len(headers)
            col_widths = [w] * len(headers)
        row_h = font_size * 0.55 + 3  # approximate row height

        def _draw_header():
            self.set_font(_FONT, "B", font_size)
            self.set_fill_color(240, 242, 245)
            self.set_text_color(30, 30, 30)
            for i, h in enumerate(headers):
                self.cell(col_widths[i], row_h + 1, str(h), border=1,
                          fill=True, align="C")
            self.ln()

        _draw_header()
        self.set_font(_FONT, "", font_size)
        self.set_text_color(50, 50, 50)
        for row in rows:
            # Page break check — re-draw header on new page
            if self.get_y() + row_h + 2 > self.h - self.b_margin:
                self.add_page()
                _draw_header()
                self.set_font(_FONT, "", font_size)
                self.set_text_color(50, 50, 50)
            for i, val in enumerate(row):
                self.cell(col_widths[i], row_h, str(val), border=1, align="C")
            self.ln()
        self.ln(4)

    def add_wide_table(self, headers, rows):
        """Render a wide table across multiple landscape pages.

        The first column ("Component") is repeated on every page-chunk.
        Value columns are split into groups that fit the page width.
        """
        fs = 7
        comp_w = 50  # width for Component column
        # Measure the widest value string to decide column width
        self.set_font(_FONT, "", fs)
        max_val_len = 0
        for row in rows:
            for val in row[1:]:
                max_val_len = max(max_val_len, self.get_string_width(str(val)))
        val_col_w = max(max_val_len + 3, 18)  # at least 18mm, plus padding

        value_headers = headers[1:]
        value_cols_data = [[row[i + 1] for row in rows] for i in range(len(value_headers))]

        # How many value columns fit per page (landscape)?
        usable_for_vals = 277 - self.l_margin - self.r_margin - comp_w  # 277mm = A4 landscape
        cols_per_page = max(1, int(usable_for_vals / val_col_w))

        # Chunk the value columns
        n_val = len(value_headers)
        for chunk_start in range(0, n_val, cols_per_page):
            chunk_end = min(chunk_start + cols_per_page, n_val)
            chunk_headers = [headers[0]] + value_headers[chunk_start:chunk_end]
            chunk_rows = []
            for r_idx, row in enumerate(rows):
                chunk_rows.append(
                    [row[0]] + [value_cols_data[c][r_idx]
                                for c in range(chunk_start, chunk_end)]
                )
            n_chunk = len(chunk_headers)
            cw = [comp_w] + [val_col_w] * (n_chunk - 1)
            self.add_page(orientation="L")
            if chunk_start > 0:
                self.set_font(_FONT, "I", 7)
                self.set_text_color(140, 140, 140)
                self.cell(0, 5, f"(continued  \u2014  columns {chunk_start + 1}"
                          f" to {chunk_end} of {n_val})")
                self.ln(7)
            self.add_table(chunk_headers, chunk_rows, col_widths=cw,
                           font_size=fs)

    def add_chart_image(self, img_bytes):
        """Insert a PNG chart (from bytes) into the PDF, auto-sized."""
        w = self.w - self.l_margin - self.r_margin  # full usable width
        # Estimate image height (assume ~55 % of width for typical charts)
        est_h = w * 0.55
        if self.get_y() + est_h > self.h - self.b_margin:
            self.add_page()
        from tempfile import NamedTemporaryFile
        with NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(img_bytes)
            tmp.flush()
            self.image(tmp.name, x=self.l_margin, w=w)
        self.ln(6)


def generate_pdf_report(
    input_values, trafo_flag, material, land_type,
    breakdown=None,
    sensitivity_results=None, sensitivity_figures=None,
    double_results=None, double_figure=None,
    double_x_label=None, double_y_label=None,
):
    """Build a complete PDF report and return it as bytes."""
    pdf = _PDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # ---- Cover / title ----
    pdf.set_font(_FONT, "B", 28)
    pdf.set_text_color(0, 102, 255)
    pdf.ln(20)
    pdf.cell(0, 15, "Investment Cost Estimation Report", align="C")
    pdf.ln(18)
    pdf.set_font(_FONT, "", 13)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 8, "EV Charging Infrastructure Investment Analysis", align="C")
    pdf.ln(10)
    pdf.set_font(_FONT, "I", 10)
    pdf.set_text_color(140, 140, 140)
    pdf.cell(0, 7, f"Generated on {date.today().strftime('%B %d, %Y')}", align="C")
    pdf.ln(30)

    # ---- 1. Input Parameters ----
    pdf.section_title(1, "Input Parameters")

    pdf.sub_heading("General")
    pdf.key_value("Transformer present", "Yes" if trafo_flag else "No")
    pdf.key_value("Material", material)
    pdf.key_value("Land Type", land_type)
    pdf.ln(3)

    pdf.sub_heading("Numerical Parameters")
    for label, value in input_values.items():
        pdf.key_value(label, str(value))
    pdf.ln(4)

    # ---- 2. Cost Breakdown ----
    if breakdown:
        pdf.add_page()
        pdf.section_title(2, "Cost Breakdown")
        headers = ["Component", "Cost"]
        rows = []
        for comp, cost in breakdown.items():
            rows.append([comp, cost if isinstance(cost, str) else str(cost)])
        cw = [(pdf.w - 20) * 0.6, (pdf.w - 20) * 0.4]
        pdf.add_table(headers, rows, col_widths=cw)

    # ---- 3. Sensitivity Analysis ----
    section_num = 3
    if sensitivity_results:
        pdf.add_page()
        pdf.section_title(section_num, "Sensitivity Analysis")
        fig_idx = 0
        for param, value_dict in sensitivity_results.items():
            # Chart first (portrait) – skipped if kaleido unavailable
            pdf.sub_heading(f"Varying: {param}")
            if sensitivity_figures and fig_idx < len(sensitivity_figures):
                img = _fig_to_bytes(sensitivity_figures[fig_idx])
                if img:
                    pdf.add_chart_image(img)
                fig_idx += 1
            # Table in landscape (auto-paging)
            cost_components = list(next(iter(value_dict.values())).keys())
            values = list(value_dict.keys())
            headers = ["Component"] + [str(v) for v in values]
            rows = []
            for component in cost_components:
                row = [component]
                for val in values:
                    row.append(value_dict[val].get(component, "\u20ac0.00"))
                rows.append(row)
            pdf.add_wide_table(headers, rows)
        section_num = 4

    # ---- 4. Two-Parameter Sensitivity ----
    if double_results and double_figure:
        pdf.add_page(orientation="P")
        pdf.section_title(section_num, "Two-Parameter Sensitivity")
        pdf.body_text(
            f"Parameters varied: {double_x_label or '?'} (X-axis) and "
            f"{double_y_label or '?'} (colour scale)."
        )
        img = _fig_to_bytes(double_figure, width=1000, height=600)
        if img:
            pdf.add_chart_image(img)
        # summary table
        (x_lbl, y_lbl), subdict = next(iter(double_results.items()))
        headers = [x_lbl, y_lbl, "Total Cost"]
        rows = []
        for (xv, yv), comps in subdict.items():
            total = sum(
                float(str(v).replace("\u20ac", "").replace(",", ""))
                for v in comps.values()
            )
            rows.append([str(xv), str(yv), f"\u20ac{total:,.2f}"])
        cw = [(pdf.w - 20) / 3] * 3
        pdf.add_table(headers, rows, col_widths=cw)

    # ---- Methodology reference ----
    pdf.add_page(orientation="P")
    pdf.section_title(section_num + 1 if double_results else section_num, "Methodology Reference")
    pdf.body_text(
        "The cost estimation methodology used in this tool is based on the master's thesis "
        "research conducted at KTH Royal Institute of Technology in collaboration with Scania AB. "
        "For full details on assumptions, data sources, component cost models, and validation, "
        "please refer to the original report:"
    )
    pdf.ln(4)
    pdf.set_font(_FONT, "", 10)
    pdf.set_text_color(0, 102, 255)
    pdf.cell(0, 7, "Open the full report on DiVA Portal", link=REPORT_URL)
    pdf.ln(12)
    pdf.set_font(_FONT, "I", 9)
    pdf.set_text_color(140, 140, 140)
    pdf.multi_cell(0, 5,
        "Disclaimer: This tool provides estimates based on reference cost data and simplified models. "
        "Actual costs may vary depending on local regulations, supplier quotes, and site-specific conditions."
    )

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf.getvalue()


# ===================================================================
#  HERO
# ===================================================================
st.markdown('<div class="hero-title">EV Charging Infrastructure Cost Estimator</div>', unsafe_allow_html=True)
st.markdown('<div class="hero-sub">Configure your charging station, pick an analysis, and get detailed cost results instantly.</div>', unsafe_allow_html=True)
st.markdown(
    '<div style="text-align:center; margin-bottom:2rem;">'
    '<a href="https://www.diva-portal.org/smash/record.jsf?dswid=1523&pid=diva2%3A2007635&c=5&searchType=SIMPLE&language=en&query=investment+cost+charging+infrastructure+heavy+duty&af=%5B%5D&aq=%5B%5B%5D%5D&aq2=%5B%5B%5D%5D&aqe=%5B%5D&noOfRows=50&sortOrder=author_sort_asc&sortOrder2=title_sort_asc&onlyFullText=true&sf=all" '
    'target="_blank" style="margin-right:1.5rem; color:#0066FF; font-weight:600; text-decoration:none;">'
    'Full Report</a>'
    '<a href="https://github.com/davide-ferraro/ev-charging-cost-estimator" '
    'target="_blank" style="color:#0066FF; font-weight:600; text-decoration:none;">'
    'GitHub</a>'
    '</div>',
    unsafe_allow_html=True,
)


# ===================================================================
#  SECTION 1 – INFRASTRUCTURE PARAMETERS
# ===================================================================
st.markdown('<hr class="section-divider">', unsafe_allow_html=True)
st.markdown('<div class="section-heading"><span class="section-number">1</span>Infrastructure Parameters</div>', unsafe_allow_html=True)
st.markdown('<div class="section-desc">Describe the electrical setup of your planned charging station.</div>', unsafe_allow_html=True)

input_values = {}

# Transformer presence
transformer_present = st.radio(
    "Is a transformer already present in your premises?",
    ["Yes", "No"], index=0, horizontal=True,
)
trafo_flag = transformer_present == "Yes"

# --- Basic Parameters ---
st.markdown('<div class="group-label">Basic Parameters</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
basic_labels = INPUT_LABELS[:7]
cols = [col1, col2, col3]
for i, label in enumerate(basic_labels):
    with cols[i % 3]:
        input_values[label] = render_input(label)

# Requirement flags
med_req = medium_requirement(
    input_values["Number of chargers"],
    input_values["Power of chargers [kW]"],
    input_values["Grid Connection capacity [kVA]"],
    input_values["Load Power Factor"],
    trafo_flag,
    input_values["Maximum power for Low Voltage Connections in your area [kVA]"],
)

# --- Low Voltage Connection ---
st.markdown('<div class="group-label">Low Voltage Connection</div>', unsafe_allow_html=True)
if not med_req:
    st.markdown('<div class="disabled-banner">Not required with current parameters</div>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
for i, label in enumerate(INPUT_LABELS[7:9]):
    with col1 if i % 2 == 0 else col2:
        input_values[label] = render_input(label, disabled=not med_req)

hard_req = hard_requirement(
    input_values["Number of chargers"],
    input_values["Power of chargers [kW]"],
    input_values["Grid Connection capacity [kVA]"],
    input_values["Load Power Factor"],
    trafo_flag,
    input_values["Maximum power for Low Voltage Connections in your area [kVA]"],
    input_values["Available Capacity of the closest MV/LV transformer [kVA]"],
)

# --- Medium Voltage Connection ---
st.markdown('<div class="group-label">Medium Voltage Connection</div>', unsafe_allow_html=True)
if not hard_req:
    st.markdown('<div class="disabled-banner">Not required with current parameters</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
for i, label in enumerate(INPUT_LABELS[9:12]):
    with [col1, col2, col3][i % 3]:
        input_values[label] = render_input(label, disabled=not hard_req)

# --- Site Preparation ---
st.markdown('<div class="group-label">Site Preparation</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
with col1:
    input_values[INPUT_LABELS[12]] = render_input(INPUT_LABELS[12])
with col2:
    material = st.radio("Material", ["Asphalt", "Concrete"], horizontal=True)
with col3:
    land_type = st.radio(
        "Land Type",
        ["High Slope / Heavy Vegetation", "Low Slope / Light Vegetation"],
    )

# Validation
if transformer_present == "No" and input_values["Grid Connection capacity [kVA]"] > input_values["Maximum power for Low Voltage Connections in your area [kVA]"]:
    st.error(
        "Invalid configuration: you selected 'No transformer present', but your "
        "grid connection exceeds the allowed threshold without a transformer."
    )
    st.stop()


# ===================================================================
#  SECTION 2 – ANALYSIS OPTIONS
# ===================================================================
st.markdown('<hr class="section-divider">', unsafe_allow_html=True)
st.markdown('<div class="section-heading"><span class="section-number">2</span>Analysis Options</div>', unsafe_allow_html=True)
st.markdown('<div class="section-desc">Choose one or more analysis types. Each will produce its own results section below.</div>', unsafe_allow_html=True)

tab_breakdown, tab_sensitivity, tab_double = st.tabs([
    "Cost Breakdown",
    "Sensitivity Analysis",
    "Two-Parameter Sensitivity",
])

with tab_breakdown:
    st.markdown("""<div class="option-card">
    <h4>Cost Breakdown</h4>
    <p>Get a full itemized breakdown of every cost component for your exact configuration:
    chargers, cables, transformers, switchgear, installation, site preparation, and more.</p>
    </div>""", unsafe_allow_html=True)
    show_breakdown = st.checkbox("Enable cost breakdown", value=True, key="cb_breakdown")

sensitivity_ranges = {}
with tab_sensitivity:
    st.markdown("""<div class="option-card">
    <h4>Sensitivity Analysis</h4>
    <p>See how the total cost changes when you vary one parameter at a time across its full range,
    while keeping everything else fixed. Useful for identifying which inputs drive costs the most.</p>
    </div>""", unsafe_allow_html=True)
    show_sensitivity = st.checkbox("Enable sensitivity analysis", key="cb_sens")
    if show_sensitivity:
        selected_params = st.multiselect("Select parameters to vary", INPUT_LABELS)
        for label in selected_params:
            minv, maxv, step, default, is_float = INPUT_DICT[label]
            c1, c2 = st.columns(2)
            if is_float:
                s_min = c1.number_input(f"{label} \u2013 Min", min_value=float(minv), max_value=float(maxv), value=float(minv), step=float(step), key=f"smin_{label}")
                s_max = c2.number_input(f"{label} \u2013 Max", min_value=float(minv), max_value=float(maxv), value=float(maxv), step=float(step), key=f"smax_{label}")
            else:
                s_min = c1.number_input(f"{label} \u2013 Min", min_value=int(minv), max_value=int(maxv), value=int(minv), step=int(step), key=f"smin_{label}")
                s_max = c2.number_input(f"{label} \u2013 Max", min_value=int(minv), max_value=int(maxv), value=int(maxv), step=int(step), key=f"smax_{label}")
            sensitivity_ranges[label] = (s_min, s_max)

double_x, double_y = None, None
with tab_double:
    st.markdown("""<div class="option-card">
    <h4>Two-Parameter Sensitivity</h4>
    <p>Vary two parameters simultaneously and visualize how their combined effect changes total cost.
    The <b>first parameter</b> is plotted on the X-axis, while the <b>second parameter</b> is shown
    as the color scale. Hover over any point to see the full component breakdown.</p>
    </div>""", unsafe_allow_html=True)
    show_double = st.checkbox("Enable two-parameter sensitivity", key="cb_double")
    if show_double:
        c1, c2 = st.columns(2)
        double_x = c1.selectbox("First parameter (X-axis)", ["None"] + INPUT_LABELS)
        double_y = c2.selectbox("Second parameter (color scale)", ["None"] + INPUT_LABELS)


# ===================================================================
#  CALCULATE
# ===================================================================
st.markdown('<hr class="section-divider">', unsafe_allow_html=True)

if st.button("Calculate", type="primary", use_container_width=True):

    # ===============================================================
    #  SECTION 3 – RESULTS
    # ===============================================================
    st.markdown('<hr class="section-divider">', unsafe_allow_html=True)
    results_container = st.container(border=False)
    with results_container:
        st.markdown('<span class="results-marker"></span>', unsafe_allow_html=True)
        st.markdown('<div class="section-heading"><span class="section-number">3</span>Results</div>', unsafe_allow_html=True)

        # Collectors for PDF export
        pdf_breakdown = None
        pdf_sensitivity_results = None
        pdf_sensitivity_figs = []
        pdf_double_results = None
        pdf_double_fig = None

        # --- Cost Breakdown ---
        if show_breakdown:
            st.markdown('<div class="group-label">Cost Breakdown</div>', unsafe_allow_html=True)
            breakdown = cost_breakdown_calculation(
                input_values, trafo_flag, material, land_type
            )
            pdf_breakdown = breakdown

            rows = [{"Component": comp, "Cost": cost} for comp, cost in breakdown.items()]
            df_display = pd.DataFrame(rows)
            st.dataframe(df_display, use_container_width=True, hide_index=True)

            excel_bytes = breakdown_to_excel(breakdown)
            st.download_button(
                "Export to Excel",
                data=excel_bytes,
                file_name="cost_breakdown.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # --- Sensitivity Analysis ---
        if show_sensitivity and sensitivity_ranges:
            st.markdown('<div class="group-label">Sensitivity Analysis</div>', unsafe_allow_html=True)
            results = cost_breakdown_sensitivity_calculation(
                input_values, trafo_flag, material, land_type, sensitivity_ranges
            )
            pdf_sensitivity_results = results

            for param, value_dict in results.items():
                st.markdown(f"**Varying: {param}**")
                cost_components = list(next(iter(value_dict.values())).keys())
                values = list(value_dict.keys())

                # Table
                table_rows = []
                for component in cost_components:
                    row = {"Component": component}
                    for val in values:
                        row[str(val)] = value_dict[val].get(component, "\u20ac0.00")
                    table_rows.append(row)
                st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

                excel_buf = sensitivity_to_excel(param, value_dict)
                st.download_button(
                    f"Export \u2018{param}\u2019 to Excel",
                    data=excel_buf,
                    file_name=f"sensitivity_{param[:30]}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{param}",
                )

                # Stacked bar chart
                x_labels = [str(v) for v in values]
                stacked_data = defaultdict(list)
                for val in values:
                    for component in cost_components:
                        if component != "Total":
                            cost_str = value_dict[val].get(component, "\u20ac0.00")
                            cost_val = float(str(cost_str).replace("\u20ac", "").replace(",", ""))
                            stacked_data[component].append(cost_val)

                colors = {comp: CHART_COLORS[i % len(CHART_COLORS)]
                          for i, comp in enumerate(stacked_data.keys())}
                traces = [
                    go.Bar(
                        name=comp, x=x_labels, y=costs,
                        marker=dict(color=colors.get(comp), line=dict(width=0)),
                    )
                    for comp, costs in stacked_data.items()
                ]

                fig = go.Figure(data=traces)
                fig.update_layout(
                    **PLOTLY_LAYOUT,
                    barmode='stack',
                    title=f"Cost Breakdown vs {param}",
                    xaxis_title=param,
                    yaxis_title="Cost (\u20ac)",
                    legend_title="Component",
                    bargap=0.15,
                    bargroupgap=0,
                )
                st.plotly_chart(fig, use_container_width=True)
                pdf_sensitivity_figs.append(fig)

        # --- Two-Parameter Sensitivity ---
        if show_double and double_x and double_y and double_x != "None" and double_y != "None":
            st.markdown('<div class="group-label">Two-Parameter Sensitivity</div>', unsafe_allow_html=True)
            results_d = cost_breakdown_double_sensitivity_calculation(
                input_values, trafo_flag, material, land_type, double_x, double_y
            )
            pdf_double_results = results_d

            (x_label, y_label), subdict = next(iter(results_d.items()))

            x_data, y_data, color_data, component_data = [], [], [], []
            for (x_val, y_val), components in subdict.items():
                total_cost = sum(
                    float(str(v).replace("\u20ac", "").replace(",", ""))
                    for v in components.values()
                )
                x_data.append(x_val)
                y_data.append(total_cost)
                color_data.append(y_val)
                component_data.append("<br>".join(
                    f"{k}: {v}" for k, v in components.items()
                ))

            fig_d = go.Figure(data=[go.Scatter(
                x=x_data, y=y_data, mode='markers',
                marker=dict(
                    size=12,
                    color=color_data,
                    colorscale=[[0, ELECTRIC_BLUE], [0.5, LIME_GREEN], [1, "#0a0a0a"]],
                    colorbar=dict(
                        title=dict(text=y_label, font=dict(color=TEXT_SECONDARY)),
                        tickfont=dict(color=TEXT_SECONDARY),
                        bgcolor="rgba(0,0,0,0)",
                        borderwidth=0,
                    ),
                    showscale=True,
                    line=dict(width=1, color="#ccc"),
                ),
                text=component_data,
                hoverinfo='text+x+y',
                name='Data points',
            )])
            fig_d.update_layout(
                **PLOTLY_LAYOUT,
                title=f"Total Cost: {x_label} vs {y_label}",
                xaxis_title=f"First Parameter: {x_label}",
                yaxis_title="Total Cost (\u20ac)",
                hovermode='closest',
            )
            st.plotly_chart(fig_d, use_container_width=True)
            pdf_double_fig = fig_d

        # --- Export PDF Report ---
        has_any = pdf_breakdown or pdf_sensitivity_results or pdf_double_results
        if has_any:
            st.markdown('<hr class="section-divider">', unsafe_allow_html=True)
            st.markdown('<div class="group-label">Export Full Report</div>', unsafe_allow_html=True)

            with st.spinner("Generating PDF report..."):
                pdf_bytes = generate_pdf_report(
                    input_values=input_values,
                    trafo_flag=trafo_flag,
                    material=material,
                    land_type=land_type,
                    breakdown=pdf_breakdown,
                    sensitivity_results=pdf_sensitivity_results,
                    sensitivity_figures=pdf_sensitivity_figs or None,
                    double_results=pdf_double_results,
                    double_figure=pdf_double_fig,
                    double_x_label=double_x,
                    double_y_label=double_y,
                )
            st.download_button(
                "Export Report as PDF",
                data=pdf_bytes,
                file_name="ev_charging_cost_report.pdf",
                mime="application/pdf",
                type="primary",
                key="dl_pdf",
            )
